import pandas as pd
import openpyxl
from openpyxl.utils.cell import get_column_letter
import os
import time

def process_and_insert_to_excel(main_csv_path):
    print("Starting data processing...")
    start_time = time.time()
    
    # Read the CSV file
    data_df = pd.read_csv(main_csv_path)
    
    # Define required columns for summary
    required_summary_cols = ['Flight', 'Origin ICAO', 'Departure Country', 'Destination ICAO', 
                            'Arriving Country', 'Block fuel', 'International']
    
    # Check if the dataframe has the required columns
    if all(col in data_df.columns for col in required_summary_cols):
        processed_df = data_df.copy()
        
        # Define CORSIA states list for 2024
        corsiaStates = [
            "Afghanistan", "Albania", "Antigua and Barbuda", "Armenia", "Australia", "Austria", "Azerbaijan",
            "Bahamas", "Bahrain", "Barbados", "Belgium", "Belize", "Benin", "Bosnia and Herzegovina",
            "Botswana", "Bulgaria", "Burkina Faso", "Cambodia", "Cameroon", "Canada", "Cook Islands",
            "Costa Rica", "Côte d'Ivoire", "Croatia", "Cuba", "Cyprus", "Czechia", 
            "Democratic Republic of the Congo", "Denmark", "Dominican Republic", "Ecuador", 
            "El Salvador", "Equatorial Guinea", "Estonia", "Finland", "France", "Gabon", "Gambia",
            "Georgia", "Germany", "Ghana", "Greece", "Grenada", "Guatemala", "Guyana", "Haiti", 
            "Honduras", "Hungary", "Iceland", "Indonesia", "Iraq", "Ireland", "Israel", "Italy", 
            "Jamaica", "Japan", "Kazakhstan", "Kenya", "Kiribati", "Kuwait", "Latvia", "Lithuania", 
            "Luxembourg", "Madagascar", "Malawi", "Malaysia", "Maldives", "Mali", "Malta", 
            "Marshall Islands", "Mauritius", "Mexico", "Micronesia (Federated States of)", "Monaco", 
            "Montenegro", "Namibia", "Nauru", "Netherlands", "New Zealand", "Nigeria", "North Macedonia", 
            "Norway", "Oman", "Palau", "Papua New Guinea", "Philippines", "Poland", "Portugal", 
            "Qatar", "Republic of Korea", "Republic of Moldova", "Romania", "Rwanda", "Saint Kitts and Nevis", 
            "Saint Vincent and the Grenadines", "Samoa", "San Marino", "Saudi Arabia", "Serbia", 
            "Seychelles", "Sierra Leone", "Singapore", "Slovakia", "Slovenia", "Solomon Islands", 
            "South Sudan", "Spain", "Suriname", "Sweden", "Switzerland", "Thailand", "Timor-Leste", 
            "Tonga", "Trinidad and Tobago", "Türkiye", "Tuvalu", "Uganda", "Ukraine", "United Arab Emirates", 
            "United Kingdom", "United Republic of Tanzania", "United States", "Uruguay", "Vanuatu", 
            "Zambia", "Zimbabwe"
        ]
        
        print("Filtering international flights...")
        # Filter for international flights only
        international_df = processed_df[processed_df['International'] == 'yes'].copy()
        
        # Convert Block fuel to numeric if it's not already
        international_df['Block fuel'] = pd.to_numeric(international_df['Block fuel'], errors='coerce')
        
        print("Creating country-level grouping...")
        # Group by country pairs and calculate totals
        country_grouped = international_df.groupby(['Departure Country', 'Arriving Country']).agg(
            Total_Flights=pd.NamedAgg(column='Flight', aggfunc='count'),
            Total_Block_Fuel=pd.NamedAgg(column='Block fuel', aggfunc='sum')
        ).reset_index()
        
        # Rename columns to match requirements
        country_grouped.rename(columns={
            'Total_Flights': 'Total No. of Flights',
            'Total_Block_Fuel': 'Block Fuel'
        }, inplace=True)
        
        # Add "Subject to offsetting requirements?" column based on CORSIA states
        country_grouped['Subject to offsetting requirements?'] = country_grouped.apply(
            lambda row: "yes" if row['Departure Country'] in corsiaStates and 
                                row['Arriving Country'] in corsiaStates else "no", 
            axis=1
        )
        
        print("Creating ICAO-level grouping...")
        # Group by ICAO pairs and calculate totals
        icao_grouped = international_df.groupby([
            'Departure Country', 'Origin ICAO', 'Arriving Country', 'Destination ICAO'
        ]).agg(
            Total_Flights=pd.NamedAgg(column='Flight', aggfunc='count'),
            Total_Block_Fuel=pd.NamedAgg(column='Block fuel', aggfunc='sum')
        ).reset_index()
        
        # Rename columns to match requirements
        icao_grouped.rename(columns={
            'Total_Flights': 'Total No. of Flights',
            'Total_Block_Fuel': 'Block Fuel'
        }, inplace=True)
        
        # Add "Subject to offsetting requirements?" column based on CORSIA states
        icao_grouped['Subject to offsetting requirements?'] = icao_grouped.apply(
            lambda row: "yes" if row['Departure Country'] in corsiaStates and 
                                row['Arriving Country'] in corsiaStates else "no", 
            axis=1
        )
        
        # Print summary of processed data
        print(f"\nProcessed \n{country_grouped} country pairs")
        print(f"Processed \n{icao_grouped} ICAO pairs")
        
        # Insert data into Excel
        print("Inserting data into Excel...")
        insert_data_to_excel(main_csv_path,country_grouped, icao_grouped)
        
        end_time = time.time()
        print(f"\nData successfully inserted into Excel template in {end_time - start_time:.2f} seconds.")
    else:
        missing = [col for col in required_summary_cols if col not in processed_df.columns]
        print(f"Warning: Cannot create summary dataframes. Missing columns: {missing}")

def insert_data_to_excel(main_csv_path,country_grouped, icao_grouped):
    # Make a copy of the template file to avoid modifying the original
    template_path = "template.xlsx"
    output_path = os.path.join(os.path.dirname(main_csv_path), 'template_filled.xlsx')
    
    # If the output file already exists, remove it
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except:
            print(f"Warning: Could not remove existing {output_path}, it may be open in another program.")
            output_path = "template_filled_new.xlsx"
    
    # Load the workbook
    print(f"Loading Excel template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    
    # Insert country-level data
    insert_country_data(wb, country_grouped)
    
    # Insert ICAO-level data
    insert_icao_data(wb, icao_grouped)
    
    # Save the workbook
    print(f"Saving output to: {output_path}")
    wb.save(output_path)

def insert_country_data(wb, country_grouped):
    try:
        # Get the sheet
        sheet_name = "5.1 Reporting-State Pairs"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            print(f"Warning: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            return
            
        print(f"Inserting {len(country_grouped)} rows into {sheet_name}...")
        
        # Calculate summary values for cells K16, K17, K18, K19, and I26
        # K16: Sum of (Block Fuel * 3.16) for all rows
        total_co2_emissions = (country_grouped['Block Fuel'] * 3.16).sum()
        sheet.cell(row=16, column=11).value = total_co2_emissions  # Column K = 11
        
        # K17: Sum of (Block Fuel * 3.16) where Subject to offsetting requirements == "yes"
        offsetting_mask = country_grouped['Subject to offsetting requirements?'] == "yes"
        offsetting_co2_emissions = (country_grouped.loc[offsetting_mask, 'Block Fuel'] * 3.16).sum()
        sheet.cell(row=17, column=11).value = offsetting_co2_emissions  # Column K = 11
        
        # K18: Sum of Total No. of Flights for all rows
        total_flights = country_grouped['Total No. of Flights'].sum()
        sheet.cell(row=18, column=11).value = total_flights  # Column L = 12
        
        # K19: Sum of Total No. of Flights where Subject to offsetting requirements == "yes"
        offsetting_flights = country_grouped.loc[offsetting_mask, 'Total No. of Flights'].sum()
        sheet.cell(row=19, column=11).value = offsetting_flights  # Column L = 12
        
        # I26: Sum of Block Fuel
        total_block_fuel = country_grouped['Block Fuel'].sum()
        sheet.cell(row=26, column=9).value = total_block_fuel  # Column I = 9
        
        # Start row for inserting data
        start_row = 56
        
        # Batch processing variables
        batch_size = 100
        total_rows = len(country_grouped)
        
        # Insert data row by row with batch reporting
        for idx, row in country_grouped.iterrows():
            current_row = start_row + idx
            
            # Progress reporting
            if idx % batch_size == 0:
                print(f"  Processing country data: {idx}/{total_rows} rows...")
            
            # State of departure
            sheet.cell(row=current_row, column=3).value = str(row['Departure Country'])
            
            # State of arrival
            sheet.cell(row=current_row, column=5).value = str(row['Arriving Country'])
            
            # CO2 emissions estimated with CERT?
            sheet.cell(row=current_row, column=7).value = "no"
            
            # Total No. of flights
            sheet.cell(row=current_row, column=8).value = int(row['Total No. of Flights'])
            
            # Fuel type
            sheet.cell(row=current_row, column=9).value = "Jet-A1"
            
            # Total mass of fuel (in tonnes)
            sheet.cell(row=current_row, column=10).value = float(row['Block Fuel'])
            
            # Total CO2 emissions (in tonnes) = 3.16 * Total mass of fuel
            fuel_mass = float(row['Block Fuel'])
            co2_emissions = 3.16 * fuel_mass
            sheet.cell(row=current_row, column=12).value = co2_emissions
            
            # Subject to offsetting requirements?
            sheet.cell(row=current_row, column=13).value = str(row['Subject to offsetting requirements?'])
            
        print(f"Completed inserting {len(country_grouped)} country rows.")
    except Exception as e:
        print(f"Error in insert_country_data: {str(e)}")

def insert_icao_data(wb, icao_grouped):
    try:
        # Get the sheet
        sheet_name = "5.2 Reporting-Aerodrome Pairs"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            print(f"Warning: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            return
            
        # Calculate summary values for cells M16, M17, M18, M19, and M26
        # M16: Sum of (Block Fuel * 3.16) for all rows
        total_co2_emissions = (icao_grouped['Block Fuel'] * 3.16).sum()
        sheet.cell(row=16, column=13).value = total_co2_emissions  # Column M = 13
        
        # M17: Sum of (Block Fuel * 3.16) where Subject to offsetting requirements == "yes"
        offsetting_mask = icao_grouped['Subject to offsetting requirements?'] == "yes"
        offsetting_co2_emissions = (icao_grouped.loc[offsetting_mask, 'Block Fuel'] * 3.16).sum()
        sheet.cell(row=17, column=13).value = offsetting_co2_emissions  # Column M = 13
        
        # M18: Sum of Total No. of Flights for all rows
        total_flights = icao_grouped['Total No. of Flights'].sum()
        sheet.cell(row=18, column=13).value = total_flights  # Column M = 13
        
        # M19: Sum of Total No. of Flights where Subject to offsetting requirements == "yes"
        offsetting_flights = icao_grouped.loc[offsetting_mask, 'Total No. of Flights'].sum()
        sheet.cell(row=19, column=13).value = offsetting_flights  # Column M = 13
        
        # H26: Sum of Block Fuel
        total_block_fuel = icao_grouped['Block Fuel'].sum()
        sheet.cell(row=26, column=8).value = total_block_fuel  # Column H = 8
        
        
        
        
        print(f"Inserting {len(icao_grouped)} rows into {sheet_name}...")
        
        # Start row for inserting data
        start_row = 57
        
        # Batch processing variables
        batch_size = 100
        total_rows = len(icao_grouped)
        
        # Insert data row by row with batch reporting
        for idx, row in icao_grouped.iterrows():
            current_row = start_row + idx
            
            # Progress reporting
            if idx % batch_size == 0:
                print(f"  Processing ICAO data: {idx}/{total_rows} rows...")
            
            # Departure ICAO airport code
            sheet.cell(row=current_row, column=3).value = str(row['Origin ICAO'])
            
            # Departure State
            sheet.cell(row=current_row, column=4).value = str(row['Departure Country'])
            
            # Arrival ICAO airport code
            sheet.cell(row=current_row, column=6).value = str(row['Destination ICAO'])
            
            # Arrival State
            sheet.cell(row=current_row, column=7).value = str(row['Arriving Country'])
            
            # CO2 emissions estimated with CERT?
            sheet.cell(row=current_row, column=9).value = "no"
            
            # Total No. of flights
            sheet.cell(row=current_row, column=10).value = int(row['Total No. of Flights'])
            
            # Fuel type
            sheet.cell(row=current_row, column=11).value = "Jet-A1"
            
            # Total mass of fuel (in tonnes)
            sheet.cell(row=current_row, column=12).value = float(row['Block Fuel'])
            
            # Total CO2 emissions (in tonnes) = 3.16 * Total mass of fuel
            fuel_mass = float(row['Block Fuel'])
            co2_emissions = 3.16 * fuel_mass
            sheet.cell(row=current_row, column=14).value = co2_emissions
            
            # Subject to offsetting requirements?
            sheet.cell(row=current_row, column=15).value = str(row['Subject to offsetting requirements?'])
            
        print(f"Completed inserting {len(icao_grouped)} ICAO rows.")
    except Exception as e:
        print(f"Error in insert_icao_data: {str(e)}")




def build_report(output_path,flight_starts_with):
    # Now load the processed file to add the Block fuel calculation and create summaries
    try:
        processed_df = pd.read_csv(output_path, encoding='utf-8')


        rows_to_delete = [] # Rows that are needed to be deleted

        # 6. FLIGHT VALIDATION
        print("Validating flight numbers...")
        if 'Flight' in processed_df.columns:
            for idx, row in processed_df.iterrows():
                if not pd.isna(row['Flight']):
                    flight_str = str(row['Flight'])
                    if flight_starts_with and not flight_str.startswith(flight_starts_with):
                        # Mark for deletion instead of error
                        if idx not in rows_to_delete:
                            rows_to_delete.append(idx)

        # Process the dataframe: remove rows marked for deletion
        if rows_to_delete:
            print(f"Removing {len(rows_to_delete)} rows that don't match flight prefix criteria")
            processed_df = processed_df.drop(rows_to_delete)

        # Now add Block fuel calculation on the processed data
        if "Block off Fuel" in processed_df.columns and "Block on Fuel" in processed_df.columns:
            # Convert columns to numeric, coercing errors to NaN
            processed_df["Block off Fuel"] = pd.to_numeric(processed_df["Block off Fuel"], errors='coerce')
            processed_df["Block on Fuel"] = pd.to_numeric(processed_df["Block on Fuel"], errors='coerce')
            
            # Calculate Block fuel as Block off Fuel - Block on Fuel
            processed_df["Block fuel"] = processed_df["Block off Fuel"] - processed_df["Block on Fuel"]
            print("Created or updated 'Block fuel' column based on the difference between Block off Fuel and Block on Fuel")
        else:
            print("Warning: Could not create 'Block fuel' column - required columns missing")
        
        # Add Departure Country and Arriving Country using airports.csv
        try:
            # Load the airports.csv file
            airports_df = pd.read_csv('airports.csv', encoding='utf-8')
            
            # Create a mapping dictionary from ICAO_Code to ICAO Member State
            icao_to_country = dict(zip(airports_df['ICAO_Code'], airports_df['ICAO Member State']))
            
            # Map Origin ICAO to Departure Country
            if 'Origin ICAO' in processed_df.columns:
                processed_df['Departure Country'] = processed_df['Origin ICAO'].map(icao_to_country)
                print("Added 'Departure Country' column based on Origin ICAO mapping")
            else:
                print("Warning: 'Origin ICAO' column not found in processed data")
            
            # Map Destination ICAO to Arriving Country
            if 'Destination ICAO' in processed_df.columns:
                processed_df['Arriving Country'] = processed_df['Destination ICAO'].map(icao_to_country)
                print("Added 'Arriving Country' column based on Destination ICAO mapping")
            else:
                print("Warning: 'Destination ICAO' column not found in processed data")
                
        except FileNotFoundError:
            print("Warning: airports.csv file not found. Cannot add country information.")
        except KeyError as e:
            print(f"Warning: Required column missing in airports.csv: {e}")
        except Exception as e:
            print(f"Warning: Error processing airports data: {str(e)}")
        
        # 4. Add International column
        if 'International' not in processed_df.columns:
            if all(col in processed_df.columns for col in ['Departure Country', 'Arriving Country']):
                # Get the index position of "Arriving Country"
                arriving_country_idx = processed_df.columns.get_loc('Arriving Country')
                
                # Create the new column values
                international_values = []
                for idx, row in processed_df.iterrows():
                    # International flight has different departure and arrival countries
                    if pd.notna(row['Departure Country']) and pd.notna(row['Arriving Country']):
                        if row['Departure Country'] != row['Arriving Country']:
                            international_values.append("yes")
                        else:
                            international_values.append("no")
                    else:
                        # Handle cases where country information is missing
                        international_values.append("unknown")
                
                # Insert the new column after "Arriving Country"
                processed_df.insert(arriving_country_idx + 1, 'International', international_values)
                print("Added 'International' column based on country comparison")
            else:
                print("Warning: Cannot add 'International' column because required columns are missing.")
        else:
            print("Warning: 'International' column already exists in processed data")
        # Save the updated processed file
        processed_df.to_csv(output_path, index=False, encoding='utf-8')
        
        process_and_insert_to_excel(output_path)
            
    except Exception as e:
        print(f"An error occurred during secondary processing: {str(e)}")
    
    print(f"\nProcessing complete.")